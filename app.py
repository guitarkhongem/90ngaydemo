import streamlit as st
import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Border, Side
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils.dataframe import dataframe_to_rows
import os
import re
from copy import copy
import logging
import io
import zipfile
import tempfile
from typing import List, Dict, Set, Optional, Any

# --- CẤU HÌNH LOGGING ---
logging.basicConfig(
    filename='full_workflow_streamlit.log',
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)

# --- CẤU HÌNH CÔNG CỤ 1: SAO CHÉP & ÁNH XẠ ---
TOOL1_COLUMN_MAPPING: Dict[str, str] = {
    'A': 'T', 'B': 'U', 'C': 'Y', 'D': 'C', 'E': 'H',
    'F': 'I', 'G': 'X', 'I': 'K', 'N': 'AX'  # Đổi từ 'AY' sang 'AX'
}
TOOL1_START_ROW_DESTINATION: int = 7
TOOL1_TEMPLATE_FILE_PATH: str = "templates/PL3-01-CV2071-QLĐĐ (Cap nhat).xlsx"
TOOL1_DESTINATION_FILE_NAME: str = "PL3-01-CV2071-QLĐĐ (Cap nhat).xlsx"

# --- CẤU HÌNH CÔNG CỤ 2: LÀM SẠCH & TÁCH FILE ---
STEP1_CHECK_COLS: List[str] = ["D", "E", "F", "I", "J", "L", "M", "R", "S", "T", "U"]
STEP1_START_ROW: int = 5
STEP1_YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
STEP1_EMPTY_FILL = PatternFill(fill_type=None)
STEP2_TARGET_COL: str = "G"
STEP2_START_ROW: int = 5
STEP2_EMPTY_FILL = PatternFill(fill_type=None)

# --- CÁC HÀM HELPER ---
def helper_copy_cell_format(src_cell, tgt_cell):
    """Sao chép định dạng từ cell nguồn sang cell đích."""
    if src_cell.has_style:
        tgt_cell.font = copy(src_cell.font)
        tgt_cell.border = copy(src_cell.border)
        tgt_cell.fill = copy(src_cell.fill)
        tgt_cell.number_format = copy(src_cell.number_format)
        tgt_cell.protection = copy(src_cell.protection)
        tgt_cell.alignment = copy(src_cell.alignment)

def helper_copy_rows_with_style(src_ws, tgt_ws, max_row=3):
    """Copy N hàng đầu tiên (giá trị + định dạng + merge + độ rộng cột)."""
    for row_idx in range(1, max_row + 1):
        for col_idx, src_cell in enumerate(src_ws[row_idx], start=1):
            tgt_cell = tgt_ws.cell(row=row_idx, column=col_idx, value=src_cell.value)
            helper_copy_cell_format(src_cell, tgt_cell)

    for col_letter, dim in src_ws.column_dimensions.items():
        if dim.width:
            tgt_ws.column_dimensions[col_letter].width = dim.width

    for merged_range in src_ws.merged_cells.ranges:
        if merged_range.min_row <= max_row:
            tgt_ws.merge_cells(str(merged_range))

def helper_normalize_value(val: Any) -> Any:
    """Chuẩn hóa giá trị: chuyển về str, loại bỏ khoảng trắng thừa, xử lý NaN."""
    if pd.isna(val) or val is None:
        return np.nan
    str_val = str(val).strip()
    str_val = re.sub(r'\s+', ' ', str_val)
    return str_val.lower() if str_val else np.nan

def helper_group_columns_openpyxl(ws):
    """Group các cột bằng openpyxl (An toàn cho môi trường online)."""
    try:
        for col in ws.column_dimensions:
            dim = ws.column_dimensions[col]
            if dim.outline_level > 0:
                dim.outline_level = 0
                dim.collapsed = False
        
        ranges_to_group = [("B", "C"), ("G", "H"), ("K", "K"), ("N", "Q"), ("W", "AX")]  # Đổi 'AY' thành 'AX'
        for start_col, end_col in ranges_to_group:
            start_idx = column_index_from_string(start_col)
            end_idx = column_index_from_string(end_col)
            for c_idx in range(start_idx, end_idx + 1):
                col_letter = get_column_letter(c_idx)
                if col_letter in ws.column_dimensions:
                    ws.column_dimensions[col_letter].outline_level = 1
        logging.info("✅ Group cột thành công bằng openpyxl")
    except Exception as e:
        logging.warning(f"⚠️ Không thể group cột bằng openpyxl: {e}")

def helper_calculate_column_width(ws):
    """Tính độ rộng cột thủ công cho openpyxl."""
    for col in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = get_column_letter(col)
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    cell_len = len(str(cell.value))
                    max_length = max(max_length, cell_len)
            except:
                pass
        adjusted_width = min(max(max_length + 2, 8), 60)
        ws.column_dimensions[column_letter].width = adjusted_width

def helper_get_safe_filepath(output_folder: str, name: str) -> str:
    """Tạo tên tệp an toàn, tránh ghi đè."""
    counter = 1
    safe_path = os.path.join(output_folder, f"{name}.xlsx")
    while os.path.exists(safe_path):
        safe_path = os.path.join(output_folder, f"{name}_{counter}.xlsx")
        counter += 1
    return safe_path

def helper_cell_has_bg(c):
    """Kiểm tra cell có màu nền hay không."""
    try:
        fg = getattr(c.fill, 'fgColor', None)
        if fg is None:
            return False
        rgb = getattr(fg, 'rgb', None)
        if rgb:
            ru = str(rgb).upper()
            if ru in ('00000000', '00000000FF', 'FFFFFFFF', '00FFFFFF', 'FF000000'):
                return False
            if ru[-6:] in ('000000', 'FFFFFF'):
                return False
            return True
        indexed = getattr(fg, 'indexed', None)
        if indexed is not None:
            return indexed != 0
        theme = getattr(fg, 'theme', None)
        if theme is not None:
            return True
    except Exception:
        return False
    return False

# --- CÁC HÀM CHO CÔNG CỤ 1 ---
def get_sheet_names_from_buffer(file_buffer: io.BytesIO) -> List[str]:
    """Đọc tên các sheet từ một buffer file Excel."""
    try:
        original_position = file_buffer.tell()
        file_buffer.seek(0)
        wb = load_workbook(file_buffer, read_only=True)
        sheet_names = wb.sheetnames
        file_buffer.seek(original_position)
        wb.close()
        return sheet_names
    except Exception as e:
        st.error(f"Không thể đọc sheet từ file: {e}")
        return []

def get_sheet_names_from_path(file_path: str) -> List[str]:
    """Đọc tên các sheet từ file Excel theo đường dẫn."""
    try:
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception as e:
        st.error(f"Không thể đọc sheet từ file mẫu: {e}")
        return []

def tool1_transform_and_copy(source_buffer, source_sheet, dest_sheet, progress_bar, status_label):
    """
    Sao chép và ánh xạ dữ liệu từ file nguồn sang file đích dựa trên file mẫu cố định.
    Áp viền cho toàn bộ vùng A:AX trong các hàng dữ liệu.
    """
    try:
        # 1. Đọc dữ liệu nguồn
        status_label.info("Đang đọc dữ liệu từ file nguồn...")
        source_cols_letters_list = list(TOOL1_COLUMN_MAPPING.keys())  # ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'I', 'N']
        source_cols_str = ",".join(source_cols_letters_list)
        
        df_source = pd.read_excel(
            source_buffer,
            sheet_name=source_sheet,
            header=None,
            skiprows=2,
            usecols=source_cols_str,
            engine='openpyxl'
        )
        
        # Kiểm tra số cột đọc được
        if len(df_source.columns) != len(source_cols_letters_list):
            st.error(f"Lỗi đọc cột: Đọc được {len(df_source.columns)} cột, nhưng mong đợi {len(source_cols_letters_list)} cột ({source_cols_str}).")
            logging.error(f"Lỗi đọc cột: Đã đọc {len(df_source.columns)} cột, mong đợi {source_cols_letters_list}")
            return None

        # Gán tên cột theo thứ tự trong TOOL1_COLUMN_MAPPING
        df_source.columns = source_cols_letters_list
        progress_bar.progress(20)

        # 2. Mở file mẫu
        status_label.info("Đang mở file mẫu...")
        if not os.path.exists(TOOL1_TEMPLATE_FILE_PATH):
            st.error(f"Lỗi: Không tìm thấy file mẫu tại '{TOOL1_TEMPLATE_FILE_PATH}'.")
            logging.error(f"Không tìm thấy file mẫu tại {TOOL1_TEMPLATE_FILE_PATH}")
            return None
        wb_dest = load_workbook(TOOL1_TEMPLATE_FILE_PATH)
        if dest_sheet not in wb_dest.sheetnames:
            st.error(f"Lỗi: Không tìm thấy sheet '{dest_sheet}' trong file mẫu.")
            logging.error(f"Sheet '{dest_sheet}' không tồn tại trong file mẫu")
            wb_dest.close()
            return None
        ws_dest = wb_dest[dest_sheet]

        # Kiểm tra và mở rộng số cột nếu cần (đến cột AX = 50)
        max_required_col = max(column_index_from_string(col) for col in TOOL1_COLUMN_MAPPING.values())  # 50 cho AX
        if ws_dest.max_column < max_required_col:
            status_label.info(f"Sheet đích chỉ có {ws_dest.max_column} cột, đang mở rộng đến {max_required_col} cột...")
            for col_idx in range(ws_dest.max_column + 1, max_required_col + 1):
                col_letter = get_column_letter(col_idx)
                ws_dest[f"{col_letter}1"] = ""  # Thêm ô trống để mở rộng cột
            logging.info(f"Đã mở rộng sheet đích lên {ws_dest.max_column} cột")
            if ws_dest.max_column < max_required_col:
                st.error(f"Lỗi: Không thể mở rộng sheet đích đến cột AX (cột {max_required_col}).")
                logging.error(f"Không thể mở rộng sheet đích đến cột {max_required_col}")
                wb_dest.close()
                return None
        progress_bar.progress(40)

        # 3. Ghi dữ liệu
        status_label.info("Đang sao chép dữ liệu...")
        total_rows_to_write = len(df_source)
        
        for source_col, dest_col in TOOL1_COLUMN_MAPPING.items():
            col_index_dest = column_index_from_string(dest_col)
            data_series = df_source[source_col]  # Lấy dữ liệu từ cột nguồn
            for j, value in enumerate(data_series, start=TOOL1_START_ROW_DESTINATION):
                cell_value = None if pd.isna(value) else value
                ws_dest.cell(row=j, column=col_index_dest, value=cell_value)
        
        progress_bar.progress(80)

        # 4. Kẻ viền cho vùng dữ liệu thực tế (A → AX)
        status_label.info("Đang kẻ viền cho vùng dữ liệu mới...")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        start_row = TOOL1_START_ROW_DESTINATION
        end_row = start_row + total_rows_to_write - 1
        start_col, end_col = 1, 50  # A:AX

        for row in ws_dest.iter_rows(min_row=start_row, max_row=end_row,
                                    min_col=start_col, max_col=end_col):
            for cell in row:
                cell.border = thin_border
        progress_bar.progress(95)

        # 5. Lưu kết quả vào buffer
        status_label.info("Đang lưu kết quả...")
        output_buffer = io.BytesIO()
        wb_dest.save(output_buffer)
        output_buffer.seek(0)
        progress_bar.progress(100)
        wb_dest.close()
        return output_buffer

    except Exception as e:
        st.error(f"Đã xảy ra lỗi trong quá trình xử lý Công cụ 1: {e}")
        logging.error(f"Lỗi Công cụ 1: {e}", exc_info=True)
        return None

# --- CÁC HÀM CHO CÔNG CỤ 2 ---
def run_step_1_process(wb, sheet_name, master_progress_bar, master_status_label, base_percent, step_budget):
    """Bước 1: Làm sạch và phân loại dữ liệu, tạo Nhóm 1 và Nhóm 2."""
    def update_progress_step1(local_percent, step_text=None):
        if step_text:
            master_status_label.info(f"Bước 1: {step_text} ({local_percent:.0f}%)")
        master_percent = base_percent + (local_percent / 100) * step_budget
        master_progress_bar.progress(int(master_percent))
    
    try:
        if sheet_name not in wb.sheetnames:
            st.error(f"Lỗi Bước 1: Không tìm thấy sheet '{sheet_name}'.")
            return None
        ws = wb[sheet_name]

        last_row = ws.max_row
        while last_row > 1 and ws[f"A{last_row}"].value in (None, ""):
            last_row -= 1
        
        update_progress_step1(0, "Đang tìm hàng trống...")
        rows_to_color = set()
        total_check_rows = last_row - STEP1_START_ROW + 1
        
        for i, row_idx in enumerate(range(STEP1_START_ROW, last_row + 1)):
            for col in STEP1_CHECK_COLS:
                cell_value = ws[f"{col}{row_idx}"].value
                if cell_value is None or str(cell_value).strip() == "":
                    rows_to_color.add(row_idx)
                    break
            if i % 100 == 0:
                update_progress_step1((i / max(total_check_rows, 1)) * 10, "Đang tìm hàng trống...")

        update_progress_step1(10, "Đang xoá màu cũ...")
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=last_row), start=1):
            for cell in row:
                cell.fill = STEP1_EMPTY_FILL
            if i % 50 == 0:
                percent = 10 + (i / last_row) * 20
                update_progress_step1(min(percent, 30), "Đang xoá màu cũ...")
        
        update_progress_step1(30, "Đang tô vàng...")
        for idx, row_idx in enumerate(rows_to_color, start=1):
            for cell in ws[row_idx]:
                cell.fill = STEP1_YELLOW_FILL
            if idx % 50 == 0:
                percent = 30 + (idx / max(len(rows_to_color), 1)) * 10
                update_progress_step1(min(percent, 40), "Đang tô vàng hàng trống...")

        update_progress_step1(40, "Đang xuất Nhóm 1...")
        ws_src = wb[sheet_name]
        last_col = ws_src.max_column

        def copy_rows_step1(title, condition_fn, start_percent, end_percent):
            if title in wb.sheetnames:
                wb.remove(wb[title])
            ws_dst = wb.create_sheet(title)
            for r in range(1, 5):
                for c in range(1, last_col + 1):
                    src = ws_src.cell(row=r, column=c)
                    dst = ws_dst.cell(row=r, column=c)
                    dst.value = src.value
                    if src.has_style:
                        helper_copy_cell_format(src, dst)
            next_row = 5
            total_data_rows = last_row - 4
            for i, r in enumerate(range(5, last_row + 1), start=1):
                if condition_fn(r):
                    for c in range(1, last_col + 1):
                        src = ws_src.cell(row=r, column=c)
                        dst = ws_dst.cell(row=next_row, column=c)
                        dst.value = src.value
                        if src.has_style:
                            helper_copy_cell_format(src, dst)
                    next_row += 1
                if i % 20 == 0:
                    progress = start_percent + (i / max(total_data_rows, 1)) * (end_percent - start_percent)
                    update_progress_step1(min(progress, end_percent), f"Đang xử lý {title}...")
            
            helper_calculate_column_width(ws_dst)

        copy_rows_step1("Nhóm 1", lambda r_idx: r_idx not in rows_to_color, 40, 70)
        copy_rows_step1("Nhóm 2", lambda r_idx: r_idx in rows_to_color, 70, 99)
        
        update_progress_step1(100, "Hoàn tất Bước 1!")
        return wb

    except Exception as e:
        st.error(f"Lỗi nghiêm trọng (Bước 1): {e}")
        logging.error(f"Lỗi Bước 1: {e}")
        return None

def run_step_2_clear_fill(wb, master_progress_bar, master_status_label, base_percent, step_budget):
    """Bước 2: Xóa màu nền các hàng có giá trị trong cột G."""
    TARGET_SHEET = "Nhóm 2"
    
    try:
        logging.info(f"Bước 2: Bắt đầu xử lý sheet {TARGET_SHEET}")
        if TARGET_SHEET not in wb.sheetnames:
            st.error(f"Lỗi (Bước 2): Không tìm thấy sheet '{TARGET_SHEET}' để xử lý.")
            return None
        ws = wb[TARGET_SHEET]
        last_row = ws.max_row
        while last_row > 1 and ws[f"A{last_row}"].value in (None, ""):
            last_row -= 1
        rows_changed = 0

        for row_idx in range(STEP2_START_ROW, last_row + 1):
            cell_g = ws[f"{STEP2_TARGET_COL}{row_idx}"]
            is_blank = (cell_g.value is None or str(cell_g.value).strip() == "")
            if not is_blank:
                for cell_in_row in ws[row_idx]:
                    cell_in_row.fill = STEP2_EMPTY_FILL
                rows_changed += 1
            if row_idx % 50 == 0:
                local_percent = (row_idx / max(last_row, 1)) * 100
                master_status_label.info(f"Bước 2: Đang xoá màu cột G... ({local_percent:.0f}%)")
                master_percent = base_percent + (local_percent / 100) * step_budget
                master_progress_bar.progress(int(master_percent))

        master_progress_bar.progress(int(base_percent + step_budget))
        logging.info(f"Bước 2: Hoàn tất, đã xoá màu {rows_changed} hàng.")
        return wb
    except Exception as e:
        st.error(f"Lỗi nghiêm trọng (Bước 2): {e}")
        logging.error(f"Lỗi Bước 2: {e}")
        return None

def run_step_3_split_by_color(wb, master_progress_bar, master_status_label, base_percent, step_budget):
    """Bước 3: Phân loại Nhóm 2 thành Nhóm 2_TC và Nhóm 2_GDC."""
    TARGET_SHEET = "Nhóm 2"
    
    try:
        logging.info(f"Bước 3: Bắt đầu xử lý sheet {TARGET_SHEET}")
        if TARGET_SHEET not in wb.sheetnames:
            st.error(f"Lỗi (Bước 3): Không tìm thấy sheet '{TARGET_SHEET}' để xử lý.")
            return None
        ws_src = wb[TARGET_SHEET]
        last_row = ws_src.max_row
        last_col = ws_src.max_column

        def copy_rows_step3(condition_fn, title):
            if title in wb.sheetnames:
                wb.remove(wb[title])
            ws_dst = wb.create_sheet(title)
            for row in range(1, 5):
                for col in range(1, last_col + 1):
                    cell_src = ws_src.cell(row=row, column=col)
                    cell_dst = ws_dst.cell(row=row, column=col)
                    cell_dst.value = cell_src.value
                    if cell_src.has_style:
                        helper_copy_cell_format(cell_src, cell_dst)
            next_row = 5
            for row in range(5, last_row + 1):
                cell = ws_src.cell(row=row, column=1)
                if condition_fn(cell):
                    for col in range(1, last_col + 1):
                        cell_src = ws_src.cell(row=row, column=col)
                        cell_dst = ws_dst.cell(row=next_row, column=col)
                        cell_dst.value = cell_src.value
                        if cell_src.has_style:
                            helper_copy_cell_format(cell_src, cell_dst)
                    next_row += 1
            
            helper_calculate_column_width(ws_dst)

        total_steps = 2 * (last_row - 4)
        processed = 0

        def update_progress_step3(add, message):
            nonlocal processed
            processed += add
            local_percent = (processed / max(total_steps, 1)) * 100
            master_status_label.info(f"Bước 3: {message} ({local_percent:.0f}%)")
            master_percent = base_percent + (local_percent / 100) * step_budget
            master_progress_bar.progress(int(master_percent))

        copy_rows_step3(lambda c: not helper_cell_has_bg(c), "Nhóm 2_TC")
        update_progress_step3(last_row - 4, "Đang xuất Nhóm 2_TC (không màu)...")

        copy_rows_step3(lambda c: helper_cell_has_bg(c), "Nhóm 2_GDC")
        update_progress_step3(last_row - 4, "Đang xuất Nhóm 2_GDC (có màu)...")

        master_progress_bar.progress(int(base_percent + step_budget))
        logging.info("Bước 3: Hoàn tất, đã tạo 'Nhóm 2_TC' và 'Nhóm 2_GDC'.")
        return wb
    except Exception as e:
        st.error(f"Lỗi nghiêm trọng (Bước 3): {e}")
        logging.error(f"Lỗi Bước 3: {e}")
        return None

def run_step_4_split_files(step4_data_buffer, main_processed_buffer, main_processed_filename, 
                          master_progress_bar, master_status_label, base_percent, step_budget):
    """Bước 4: Tách file Nhóm 2_GDC theo cột T và nén thành ZIP."""
    DATA_SHEET = "Nhóm 2_GDC"
    TEMPLATE_SHEET = "TongHop"
    FILTER_COLUMN = "T"
    START_ROW = 5
    
    try:
        logging.info("Bước 4: Bắt đầu xử lý tách file")
        wb_openpyxl = load_workbook(step4_data_buffer, data_only=True)
        if TEMPLATE_SHEET not in wb_openpyxl.sheetnames:
            st.error("Lỗi (Bước 4): Không tìm thấy sheet mẫu 'TongHop'!")
            return None
        if DATA_SHEET not in wb_openpyxl.sheetnames:
            st.error("Lỗi (Bước 4): Không tìm thấy sheet dữ liệu 'Nhóm 2_GDC'!")
            return None
        tonghop_ws = wb_openpyxl[TEMPLATE_SHEET]
        
        step4_data_buffer.seek(0)
        df = pd.read_excel(step4_data_buffer, sheet_name=DATA_SHEET, header=None)
        logging.info("Đã tải thành công template và data từ buffer")

        col_index = column_index_from_string(FILTER_COLUMN) - 1
        start_row_index = START_ROW - 1
        if col_index >= len(df.columns):
            st.error(f"Lỗi (Bước 4): Cột lọc '{FILTER_COLUMN}' không tồn tại!")
            return None
        
        data_col_raw = df.iloc[start_row_index:, col_index]
        data_col = data_col_raw.apply(helper_normalize_value)
        unique_normalized = data_col.dropna().unique().tolist()
        if data_col.isnull().any():
            unique_normalized.append("BLANK")

        total = len(unique_normalized)
        master_status_label.info(f"Bước 4: Chuẩn bị tách {total} file con...")

        with tempfile.TemporaryDirectory() as tmpdir:
            logging.info(f"Đã tạo thư mục tạm: {tmpdir}")
            try:
                main_file_path = os.path.join(tmpdir, main_processed_filename)
                with open(main_file_path, 'wb') as f:
                    f.write(main_processed_buffer.getbuffer())
                logging.info(f"Đã lưu file chính vào: {main_file_path}")
            except Exception as e_save_main:
                logging.warning(f"Không thể lưu file chính vào zip: {e_save_main}")

            for i, norm_value in enumerate(unique_normalized, start=1):
                if norm_value == "BLANK":
                    mask = data_col.isnull()
                else:
                    mask = data_col == norm_value
                filtered = df.iloc[start_row_index:][mask]
                if filtered.empty:
                    continue

                new_wb = Workbook()
                new_ws = new_wb.active
                new_ws.title = "DuLieuLoc"
                helper_copy_rows_with_style(tonghop_ws, new_ws, max_row=3)
                for r_idx, row in enumerate(dataframe_to_rows(filtered, index=False, header=False), start=4):
                    for c_idx, value_cell in enumerate(row, start=1):
                        new_ws.cell(row=r_idx, column=c_idx, value=value_cell)
                
                safe_name = "BLANK" if norm_value == "BLANK" else re.sub(r'[\\/*?:<>|"\t\n\r]+', "_", str(norm_value).strip())[:50]
                output_path = helper_get_safe_filepath(tmpdir, safe_name)
                
                try:
                    helper_group_columns_openpyxl(new_ws)
                    helper_calculate_column_width(new_ws)
                    new_wb.save(output_path)
                except Exception as e_openpyxl:
                    logging.error(f"Lỗi openpyxl khi xử lý {output_path}: {e_openpyxl}")
                finally:
                    new_wb.close()

                local_percent = (i / total) * 100
                master_status_label.info(f"Bước 4: Đang tách file {i}/{total} ({local_percent:.0f}%)")
                master_percent = base_percent + (local_percent / 100) * step_budget
                master_progress_bar.progress(int(master_percent))
            
            master_status_label.info("Bước 4: Đang nén file ZIP...")
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_f:
                for root, _, files in os.walk(tmpdir):
                    for file in files:
                        zip_f.write(os.path.join(root, file), arcname=file)
            
            zip_buffer.seek(0)
            master_progress_bar.progress(int(base_percent + step_budget))
            logging.info("Đã tạo ZIP buffer thành công.")
            
            wb_openpyxl.close()
            return zip_buffer

    except Exception as e:
        st.error(f"Lỗi nghiêm trọng (Bước 4): {str(e)}")
        logging.error(f"Lỗi Bước 4: {e}")
        return None
    finally:
        if 'wb_openpyxl' in locals():
            try:
                wb_openpyxl.close()
            except:
                pass

# --- GIAO DIỆN STREAMLIT CHÍNH ---
st.set_page_config(page_title="TSCopyRight", layout="wide", page_icon="🚀")

# --- SIDEBAR ---
st.sidebar.title("Hướng dẫn sử dụng")
st.sidebar.markdown("""
- **Kế hoạch số 515/KH-BCA-BNN&MT ngày 31/8/2025 của Bộ Công an và Bộ Nông nghiệp và Môi trường về việc triển khai thực hiện chiến dịch làm giàu, làm sạch cơ sở dữ liệu quốc gia về đất đai.
- **Công văn số 780/UBND-NNMT ngày 04/9/2025 của UBND tỉnh Quảng Trị về việc triển khai Kế hoạch số 515/KH-BCA-BNN&MT.
- **Công văn số 2071/QLĐĐ-TKKKTTĐĐ ngày 05/9/2025 của Cục Quản lý đất đai về việc hướng dẫn tổ chức thực hiện chiến dịch làm giàu, làm sạch cơ sở dữ liệu quốc gia về đất đai.
- **Công văn số 1730/SNNMT-ĐĐBĐVT ngày 08/9/2025 của Sở Nông nghiệp và Môi trường tỉnh Quảng Trị về việc triển khai Kế hoạch số 515/KH-BCA-BNN&MT.
- **Quyết định 1392/QĐ-UBND ngày 10/9/2025, của UBND tỉnh về việc thành lập Tổ công tác.
- **Kế hoạch số 847/KH-UBND ngày 10/9/2025, của UBND tỉnh Quảng Trị về triển khai thực hiện chiến dịch làm giàu, làm sạch cơ sở dữ liệu đất đai.
- **Công văn số 2240/QLĐĐ-TKKKTTĐĐ ngày 19/9/2025, về việc phối hợp với các đơn vị phần mềm trong thực hiện Kế hoạch số 515/KH-BCA-BNN&MT.
- **Công văn số /QLĐĐ-TKKKTTĐĐ Tháng 10 năm 2025, về tài liệu hướng dẫn bổ sung theo Công văn số 2071/QLĐĐ-TKKKTTĐĐ.
""")
st.sidebar.info("Phát triển dựa trên quy trình nghiệp vụ của Trường Sinh - SĐT 0917.750.555.")

# --- MAIN PAGE ---
st.title("Chiến Dịch Xây Dựng Cơ Sở Dữ Liệu Đất Đai")
st.header("Bộ Công cụ Hỗ trợ Dữ liệu")
st.markdown("---")

# --- TẠO 2 TAB CHO 2 CÔNG CỤ ---
tab1, tab2 = st.tabs([
    "Công cụ 1: Sao chép & Ánh xạ Dữ liệu",
    "Công cụ 2: Làm sạch & Tách file (Quy trình chính)"
])

# --- GIAO DIỆN CHO CÔNG CỤ 1 ---
with tab1:
    st.subheader("Sao chép dữ liệu từ File Nguồn sang File Mẫu")
    
    st.markdown("### Bước 1: Tải lên File Nguồn (File chứa dữ liệu)")
    source_file = st.file_uploader("Chọn File Nguồn (.xlsx, .xlsm)", type=["xlsx", "xlsm"], key="tool1_source")
    
    source_sheet = None
    dest_sheet = None

    col1, col2 = st.columns(2)
    with col1:
        if source_file:
            source_sheets = get_sheet_names_from_buffer(source_file)
            source_sheet = st.selectbox("Chọn Sheet Nguồn (để đọc):", source_sheets, key="tool1_source_sheet")
    
    with col2:
        try:
            dest_sheets = get_sheet_names_from_path(TOOL1_TEMPLATE_FILE_PATH)
            dest_sheet = st.selectbox("Chọn Sheet Đích (để ghi):", dest_sheets, key="tool1_dest_sheet")
        except Exception as e:
            st.error(f"Không thể đọc file mẫu tại '{TOOL1_TEMPLATE_FILE_PATH}'. Vui lòng kiểm tra!")
            logging.error(f"Lỗi đọc file mẫu: {e}")
            dest_sheet = None

    st.markdown("### Bước 2: Xác nhận")
    start_tool1 = st.button("Bắt đầu Sao chép & Ánh xạ", key="tool1_start")

    if start_tool1:
        if not source_file or not source_sheet or not dest_sheet:
            st.error("Vui lòng tải lên file nguồn và chọn cả hai sheet.")
        else:
            progress_bar_tool1 = st.progress(0)
            status_label_tool1 = st.empty()
            
            try:
                source_file.seek(0)
                result_buffer = tool1_transform_and_copy(
                    source_file, source_sheet, 
                    dest_sheet, 
                    progress_bar_tool1, status_label_tool1
                )
                
                if result_buffer:
                    status_label_tool1.success("✅ HOÀN TẤT!")
                    st.download_button(
                        label="Tải về File Đích đã cập nhật",
                        data=result_buffer,
                        file_name=TOOL1_DESTINATION_FILE_NAME,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                else:
                    status_label_tool1.error("Xử lý thất bại. Vui lòng kiểm tra log.")
            
            except Exception as e:
                st.error(f"Lỗi nghiêm trọng Công cụ 1: {e}")
                logging.error(f"Lỗi Streamlit Tool 1: {e}", exc_info=True)

# --- GIAO DIỆN CHO CÔNG CỤ 2 ---
with tab2:
    st.subheader("Làm sạch, Phân loại và Tách file tự động")
    
    st.markdown("### Bước 1: Tải lên File Excel")
    uploaded_file_tool2 = st.file_uploader("Chọn file Excel của bạn (.xlsx, .xlsm)", type=["xlsx", "xlsm"], key="tool2_uploader")

    if uploaded_file_tool2:
        st.markdown("---")
        st.markdown("### Bước 2: Chọn Sheet")
        try:
            uploaded_file_tool2.seek(0)
            wb_sheets = load_workbook(uploaded_file_tool2, read_only=True)
            sheet_names = wb_sheets.sheetnames
            wb_sheets.close()
            
            selected_sheet_tool2 = st.selectbox("Chọn sheet chính để xử lý:", sheet_names, 
                                               help="Đây là sheet gốc chứa dữ liệu bạn muốn lọc.", 
                                               key="tool2_sheet_select")

            st.markdown("### Bước 3: Xác nhận")
            start_button_tool2 = st.button("Bắt đầu Làm sạch & Tách file", key="tool2_start")
            st.markdown("---")

            if start_button_tool2:
                st.markdown("### Bước 4: Hoàn thành và Tải về")
                progress_bar = st.progress(0)
                status_text_area = st.empty()
                
                try:
                    status_text_area.info("Đang tải file vào bộ nhớ...")
                    uploaded_file_tool2.seek(0)
                    main_wb = load_workbook(uploaded_file_tool2)
                    
                    main_wb = run_step_1_process(main_wb, selected_sheet_tool2, progress_bar, status_text_area, 0, 25)
                    if main_wb is None: raise Exception("Bước 1 thất bại.")
                    
                    main_wb = run_step_2_clear_fill(main_wb, progress_bar, status_text_area, 25, 25)
                    if main_wb is None: raise Exception("Bước 2 thất bại.")
                    
                    main_wb = run_step_3_split_by_color(main_wb, progress_bar, status_text_area, 50, 25)
                    if main_wb is None: raise Exception("Bước 3 thất bại.")
                    
                    status_text_area.info("Đang chuẩn bị file kết quả...")
                    final_wb_buffer = io.BytesIO()
                    main_wb.save(final_wb_buffer)
                    final_wb_buffer.seek(0)
                    
                    step4_read_buffer = io.BytesIO(final_wb_buffer.read())
                    final_wb_buffer.seek(0)
                    
                    main_processed_filename = f"[Processed]_{uploaded_file_tool2.name}"
                    
                    zip_buffer = run_step_4_split_files(
                        step4_read_buffer,
                        final_wb_buffer,
                        main_processed_filename,
                        progress_bar, 
                        status_text_area, 
                        75, 
                        25
                    )
                    if zip_buffer is None: raise Exception("Bước 4 thất bại.")

                    main_wb.close()
                    
                    status_text_area.success("✅ HOÀN TẤT!")
                    progress_bar.progress(100)
                    
                    st.download_button(
                        label="🗂️ Tải về Gói Kết Quả (ZIP)",
                        data=zip_buffer,
                        file_name="KetQua_Thon.zip",
                        mime="application/zip",
                        help=f"File ZIP này chứa file Excel chính ({main_processed_filename}) VÀ tất cả các file con được tách ra từ 'Nhóm 2_GDC'."
                    )
                    
                except Exception as e:
                    st.error(f"Quy trình đã dừng do lỗi: {e}")
                    logging.error(f"Lỗi Streamlit Workflow Tool 2: {e}")

        except Exception as e:
            st.error(f"Không thể đọc file Excel. File có thể bị hỏng hoặc không đúng định dạng: {e}")
            logging.error(f"Lỗi Streamlit Tải file Tool 2: {e}")